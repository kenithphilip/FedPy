"""Offline end-to-end tests for Stage 3 evidence enrichment (no AWS).

Covers: per-requirement evidence mapping, scope propagation, indicators (overall +
per-scope), risk lists + thresholds, findings join, the augment-not-overwrite
guarantee, the consolidated document/workbook/CSV, and the scope-missing fallback.
"""

from __future__ import annotations

import json

from openpyxl import load_workbook

from pci_inventory.evidence.findings import Finding, FindingsResult
from pci_inventory.evidence.indicators import compute_indicators
from pci_inventory.evidence.loader import LoadedArtifact, load_for_evidence
from pci_inventory.evidence.mapping import build_evidence_rows
from pci_inventory.evidence.models import EvidenceThresholds, RequirementDomain
from pci_inventory.evidence.runner import (
    build_evidence_document,
    run_evidence,
)
from pci_inventory.evidence.csv_writer import write_evidence_csv
from pci_inventory.evidence.workbook import write_evidence_workbook
from pci_inventory.scope.artifact import InventoryIndex


def _scoped_doc():
    def r(arn, rid, rtype, **kw):
        base = {"arn": arn, "resource_id": rid, "resource_type": rtype, "name": kw.get("name", rid),
                "region": kw.get("region", "us-east-1"), "account_id": "111122223333"}
        base.update({k: v for k, v in kw.items() if k not in ("name", "region")})
        base.setdefault("scope", {"category": kw.get("cat", "CDE"), "confidence": "DETERMINED",
                                  "basis": ["seed"]})
        return base
    return {
        "schema_version": "1.1.0", "scope_schema_version": "1.0.0",
        "accounts_scanned": [{"account_id": "111122223333", "alias": "acme"}],
        "regions_coverage": [{"account_id": "111122223333", "region": "us-east-1", "status": "included"}],
        "scope_analysis": {"stats": {}},
        "resources": [
            r("arn:db", "pay-db", "rds:db-cluster", encryption_at_rest=True,
              encryption_in_transit=True, tls_min_version="TLSv1.2", iam_db_auth=True,
              publicly_shared=False, data_classification="chd", cat="CDE"),
            r("arn:bucket", "logs", "s3:bucket", encryption_at_rest=False, public_exposed=True,
              public_access_block=False, cat="connected-to"),
            r("arn:inst", "i-1", "ec2:instance", imdsv2_required=False,
              anti_malware_status="NOT_COLLECTABLE", cat="CDE"),
            r("arn:user", "alice", "iam:user", region="GLOBAL", mfa_enabled=False,
              access_key_age_days="420", cat="security-impacting",
              iam_policy_data={"principal_type": "iam:user",
                               "attached_managed_policies": ["arn:aws:iam::aws:policy/AdministratorAccess"],
                               "inline_policies": {}}),
            r("arn:lg", "/aws/x", "logs:log-group", logging_enabled=True, log_retention_days="90",
              cat="security-impacting"),
        ],
    }


def test_evidence_mapping_per_requirement():
    doc = _scoped_doc()
    db = doc["resources"][0]
    rows = build_evidence_rows(db)
    reqs = {row.requirement for row in rows}
    assert RequirementDomain.REQ3 in reqs  # data at rest
    assert RequirementDomain.REQ4 in reqs  # crypto in transit
    assert RequirementDomain.REQ8 in reqs  # DB IAM auth
    assert RequirementDomain.REQ12 in reqs  # inventory
    # Scope propagated onto every row.
    assert all(row.scope_category == "CDE" and row.scope_confidence == "DETERMINED" for row in rows)


def test_indicators_overall_and_by_scope():
    res = _scoped_doc()["resources"]
    ind = compute_indicators(res, EvidenceThresholds())
    assert ind.overall["encryption_at_rest_pct"] == 50.0  # db True, bucket False
    assert ind.overall["imdsv2_enforcement_pct"] == 0.0
    assert ind.overall["public_exposed_count"] == 1
    assert "CDE" in ind.by_scope and "security-impacting" in ind.by_scope
    # Risk lists.
    assert any("arn:bucket" in x for x in ind.risk_lists["unencrypted_at_rest"])
    assert any("arn:user" in x for x in ind.risk_lists["stale_credentials"])
    assert any("AdministratorAccess" in x for x in ind.risk_lists["overly_permissive_iam"])
    assert any("arn:lg" in x for x in ind.risk_lists["log_retention_below_threshold"])


def test_threshold_override_changes_stale():
    res = _scoped_doc()["resources"]
    # raise the stale threshold above the 420d key → no longer flagged
    ind = compute_indicators(res, EvidenceThresholds(stale_credential_days=500))
    assert ind.risk_lists["stale_credentials"] == []


def test_findings_join_by_arn():
    fr = FindingsResult(findings=[
        Finding("guardduty", "8", "Backdoor:EC2/C&CActivity", "arn:inst", "111122223333", "us-east-1"),
        Finding("securityhub", "HIGH", "S3.8 block public access", "arn:bucket", "111122223333", "us-east-1"),
        Finding("inspector2", "CRITICAL", "orphan finding", "arn:not-in-inventory", "111122223333", "us-east-1"),
    ])
    loaded = LoadedArtifact(InventoryIndex(_scoped_doc()), scope_missing=False, source_path=None)
    result = run_evidence(loaded, EvidenceThresholds(), session_factory=None, app_cfg=None)
    # inject findings + re-attach (run_evidence used no session so findings empty)
    result.findings = fr
    from pci_inventory.evidence.runner import self_attach_findings
    self_attach_findings(result)
    inst_rows = [r for r in result.rows if r.resource_arn == "arn:inst"]
    assert any(any("guardduty" in f for f in r.findings) for r in inst_rows)
    # orphan finding is retained in the result set even if no row matched
    assert any(f.resource_arn == "arn:not-in-inventory" for f in result.findings.findings)


def test_augment_not_overwrite():
    loaded = LoadedArtifact(InventoryIndex(_scoped_doc()), scope_missing=False, source_path=None)
    result = run_evidence(loaded, EvidenceThresholds(), session_factory=None, app_cfg=None)
    doc = build_evidence_document(loaded, result)
    db = next(r for r in doc["resources"] if r["arn"] == "arn:db")
    # Stage 1/2 fields preserved verbatim.
    assert db["encryption_at_rest"] is True
    assert db["scope"]["category"] == "CDE"
    # Stage 3 augmentation added.
    assert "evidence" in db
    assert "Req 03" in db["evidence"]["by_requirement"]
    assert doc["evidence_schema_version"] == "1.0.0"
    assert "indicators" in doc["evidence_analysis"]
    assert "requirement_mapping" in doc["evidence_analysis"]


def test_workbook_and_csv(tmp_path):
    loaded = LoadedArtifact(InventoryIndex(_scoped_doc()), scope_missing=False, source_path=None)
    result = run_evidence(loaded, EvidenceThresholds(), session_factory=None, app_cfg=None)
    xlsx = write_evidence_workbook(result, tmp_path / "ev.xlsx")
    wb = load_workbook(xlsx)
    # all 12 requirement sheets + mapping + indicators + notes
    for name in ("Req 01 NSC", "Req 03 Data at Rest", "Req 08 Identity & Auth", "Req 12 Program & Scope",
                 "PCI Requirement Mapping", "Findings & Indicators", "QSA Notes"):
        assert name in wb.sheetnames
    csv_path = write_evidence_csv(result, tmp_path / "ev.csv")
    lines = csv_path.read_text().splitlines()
    assert lines[0].startswith("requirement,sub_requirements,resource_arn")
    assert len(lines) - 1 == len(result.rows)


def test_scope_missing_fallback(tmp_path):
    # Only a Stage 1 inventory (no scope) → scope_missing banner + undetermined scope.
    inv = {"schema_version": "1.1.0",
           "accounts_scanned": [{"account_id": "1", "alias": "a"}],
           "regions_coverage": [],
           "resources": [{"arn": "arn:x", "resource_id": "x", "resource_type": "s3:bucket",
                          "name": "x", "region": "us-east-1", "account_id": "1",
                          "encryption_at_rest": True}]}
    ip = tmp_path / "inventory.json"
    ip.write_text(json.dumps(inv))
    loaded = load_for_evidence(tmp_path / "missing-scoped.json", ip)
    assert loaded.scope_missing is True
    result = run_evidence(loaded, EvidenceThresholds(), session_factory=None, app_cfg=None)
    assert any("SCOPE CONTEXT MISSING" in c for c in result.caveats)
    assert all(r.scope_category == "undetermined" for r in result.rows)


def test_indicator_disclaimer_present():
    res = _scoped_doc()["resources"]
    ind = compute_indicators(res, EvidenceThresholds())
    assert "NOT" in ind.disclaimer and "determination" in ind.disclaimer.lower()
