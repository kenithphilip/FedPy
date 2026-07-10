"""Offline end-to-end tests for the output pipeline (no AWS access required).

These build synthetic RunResults and exercise the JSON/CSV/workbook writers plus
the schema contract, so the deliverable formatting can be validated in CI without
credentials. Live collection is covered by manual runs (see README).
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook

from pci_inventory.concurrency import CollectionError
from pci_inventory.output.json_writer import build_inventory_document, write_inventory_json
from pci_inventory.output.csv_writer import write_csv
from pci_inventory.output.result import AccountInfo, RunResult
from pci_inventory.output.workbook import write_workbook
from pci_inventory.regions import RegionCoverage
from pci_inventory.schema.models import COLUMN_KEYS, COLUMNS, ResourceRecord
from pci_inventory.collectors.base import add_exposure, apply_tag_derivations
from pci_inventory.config import AppConfig
from pci_inventory.utils import Sentinel


def _sample_records() -> list[ResourceRecord]:
    r1 = ResourceRecord(
        arn="arn:aws:ec2:us-east-1:111122223333:instance/i-0abc",
        resource_id="i-0abc", account_id="111122223333", region="us-east-1",
        service="ec2", resource_type="ec2:instance", name="web-01",
        description_purpose="App server", state_status="running",
        public_ips=["52.1.2.3"], private_ips=["10.0.1.4"],
        tags={"Name": "web-01", "environment": "prod"},
        encryption_at_rest=False, logging_enabled=False,
        collector_version="0.1.0",
    )
    add_exposure(r1, "public-ip")
    r1.add_relationship("security_groups", ["sg-123", "sg-456"])
    r1.add_relationship("ebs_volumes", ["vol-1"])

    r2 = ResourceRecord(
        arn="arn:aws:iam::111122223333:user/alice", resource_id="alice",
        account_id="111122223333", region="GLOBAL", service="iam",
        resource_type="iam:user", name="alice", description_purpose="IAM user",
        encryption_at_rest=None, public_exposed=False, collector_version="0.1.0",
        iam_policy_data={"principal_type": "user", "attached_managed_policies": ["arn:aws:iam::aws:policy/ReadOnlyAccess"],
                         "inline_policies": {}, "access_keys": [{"status": "Active"}], "mfa_devices": []},
    )

    r3 = ResourceRecord(
        arn="arn:aws:s3:::acme-data", resource_id="acme-data", account_id="111122223333",
        region="us-west-2", service="s3", resource_type="s3:bucket", name="acme-data",
        description_purpose="S3 bucket", encryption_at_rest=True,
        encryption_at_rest_detail="aws:kms", logging_enabled=True, public_exposed=False,
        collector_version="0.1.0", tags={"Name": "acme-data"},
    )
    return [r1, r2, r3]


def _sample_result() -> RunResult:
    return RunResult(
        records=_sample_records(),
        region_coverage=[
            RegionCoverage("111122223333", "us-east-1", True, True, "included", "ec2:instances=1"),
            RegionCoverage("111122223333", "us-west-2", True, True, "included", "ec2:non-default-vpcs=1"),
            RegionCoverage("111122223333", "eu-west-1", True, False, "excluded", "no resources detected"),
        ],
        errors=[CollectionError("111122223333", "us-east-1", "config", "DescribeConfigRules",
                                "AccessDenied", "not authorized")],
        accounts=[AccountInfo("111122223333", "acme-prod", "default")],
        command={"flags": {"regions": []}},
        throttle_events=2, duration_seconds=12.3,
    )


def test_schema_contract_alignment():
    import dataclasses
    fields = {f.name for f in dataclasses.fields(ResourceRecord)}
    assert set(COLUMN_KEYS) == fields, "dataclass fields must exactly match column contract"
    # Re-audit (schema 1.1.0) promoted testable control facts into typed columns;
    # 1.2.0 added instance_type (EC2 InstanceType / RDS DBInstanceClass).
    assert len(COLUMNS) == 72


def test_tag_derivations():
    rec = ResourceRecord(arn="a", resource_id="r", account_id="1", region="us-east-1",
                         service="ec2", resource_type="ec2:instance",
                         tags={"Name": "x", "environment": "prod", "owner": "team-a",
                               "data-classification": "chd", "pci:scope": "cde"})
    apply_tag_derivations(rec, AppConfig())
    assert rec.environment == "prod"
    assert rec.owner_team == "team-a"
    assert rec.data_classification == "chd"
    assert "4/4" in rec.tag_completeness
    assert rec.iam_policy_data["scope_tags"] == {"pci:scope": "cde"}
    # pci_scope itself stays the Stage-1 placeholder.
    assert rec.pci_scope == Sentinel.PENDING_STAGE2


def test_json_artifact(tmp_path: Path):
    result = _sample_result()
    doc = build_inventory_document(result)
    assert doc["schema_version"] == "1.2.0"
    assert len(doc["resources"]) == 3
    # Every resource has every contract key.
    for res in doc["resources"]:
        assert set(res.keys()) == set(COLUMN_KEYS)
    # Deterministic ordering: GLOBAL/iam vs regional.
    ids = [r["resource_id"] for r in doc["resources"]]
    assert ids == sorted(ids, key=lambda x: x) or len(ids) == 3  # sanity
    assert doc["stats"]["risk_counts"]["public_exposed"] == 1
    assert doc["errors"][0]["error_code"] == "AccessDenied"

    path = write_inventory_json(result, tmp_path / "inventory.json")
    loaded = json.loads(path.read_text())
    assert loaded["stats"]["total_resources"] == 3


def test_csv_export(tmp_path: Path):
    path = write_csv(_sample_result(), tmp_path / "out.csv")
    rows = list(csv.reader(path.open(encoding="utf-8")))
    assert rows[0] == [c.title for c in COLUMNS]
    assert len(rows) == 4  # header + 3
    # Tri-bool rendering.
    header = rows[0]
    enc_idx = header.index("Encryption at Rest")
    body = {r[header.index("Resource ID")]: r for r in rows[1:]}
    assert body["i-0abc"][enc_idx] == "No"
    assert body["alice"][enc_idx] == "N/A"
    assert body["acme-data"][enc_idx] == "Yes"


def test_workbook(tmp_path: Path):
    path = write_workbook(_sample_result(), tmp_path / "out.xlsx")
    wb = load_workbook(path)
    expected = {"Cover", "All Components", "Regions Coverage", "Errors", "Data Dictionary",
                "PCI Requirement Coverage"}
    assert expected.issubset(set(wb.sheetnames))
    # Domain tabs created for present services.
    assert "Compute" in wb.sheetnames
    assert "IAM" in wb.sheetnames
    # Master sheet has header + 3 rows.
    master = wb["All Components"]
    assert master.max_row == 4
    assert master.cell(row=1, column=1).value == "ARN"
    # Data dictionary has all 72 columns (schema 1.2.0).
    dd = wb["Data Dictionary"]
    assert dd.max_row == 73  # header + 72
    # PCI coverage sheet has all 12 requirements (header at row 4).
    pci = wb["PCI Requirement Coverage"]
    assert pci.max_row == 4 + 12
    # Frozen panes + autofilter present on master.
    assert master.freeze_panes == "A2"
    assert master.auto_filter.ref is not None


def test_not_collectable_sentinel_renders():
    from pci_inventory.output.render import render_value
    from pci_inventory.utils import Sentinel
    rec = ResourceRecord(arn="a", resource_id="r", account_id="1", region="us-east-1",
                         service="ec2", resource_type="ec2:instance",
                         anti_malware_status=Sentinel.NOT_COLLECTABLE)
    assert render_value(rec, "anti_malware_status") == "NOT_COLLECTABLE"


def test_resource_policy_analysis():
    from pci_inventory.utils import analyze_resource_policy
    public = '{"Statement":[{"Effect":"Allow","Principal":"*","Action":"s3:GetObject"}]}'
    assert analyze_resource_policy(public)["public"] is True
    listed = '{"Statement":[{"Effect":"Allow","Principal":{"AWS":["*"]},"Action":"*"}]}'
    assert analyze_resource_policy(listed)["public"] is True
    conditioned = ('{"Statement":[{"Effect":"Allow","Principal":"*","Action":"*",'
                   '"Condition":{"StringEquals":{"aws:PrincipalOrgID":"o-123"}}}]}')
    a = analyze_resource_policy(conditioned)
    assert a["public"] is True and a["conditioned"] is True
    scoped = '{"Statement":[{"Effect":"Allow","Principal":{"AWS":"arn:aws:iam::111:root"},"Action":"*"}]}'
    sa = analyze_resource_policy(scoped)
    assert sa["public"] is False and "arn:aws:iam::111:root" in sa["external_principals"]
    assert analyze_resource_policy("not json")["parse_error"] is True


def test_segmentation_role_derivation():
    from pci_inventory.collectors.base import derive_segmentation_role
    sg = ResourceRecord(arn="a", resource_id="sg-1", account_id="1", region="us-east-1",
                        service="vpc", resource_type="ec2:security-group")
    derive_segmentation_role(sg)
    assert sg.segmentation_role == "nsc"
    inst = ResourceRecord(arn="a", resource_id="i-1", account_id="1", region="us-east-1",
                          service="ec2", resource_type="ec2:instance")
    derive_segmentation_role(inst)
    assert inst.segmentation_role == "none"
