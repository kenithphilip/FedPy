"""Offline end-to-end tests for Stage 2 scope analysis (no AWS).

Builds a synthetic inventory.json-shaped document with a small VPC topology and
IAM, then drives the seed resolver, reachability graph, IAM graph, and classifier
to assert correct classifications, proven paths, and a segmentation finding.
"""

from __future__ import annotations

from pci_inventory.scope.artifact import InventoryIndex
from pci_inventory.scope.classifier import ScopeClassifier
from pci_inventory.scope.gapfetch import (
    Nacl,
    NaclRule,
    NetworkData,
    Route,
    RouteTable,
)
from pci_inventory.scope.iamgraph import IamGraph
from pci_inventory.scope.models import Category, Confidence
from pci_inventory.scope.reachability import ReachabilityGraph
from pci_inventory.scope.seeds import SeedResolver, load_scope_config


def _res(arn, rid, rtype, **kw):
    base = {
        "arn": arn, "resource_id": rid, "resource_type": rtype, "account_id": "111122223333",
        "region": "us-east-1", "name": kw.pop("name", rid), "tags": kw.pop("tags", {}),
        "relationships": kw.pop("relationships", {}), "iam_policy_data": kw.pop("iam_policy_data", {}),
        "public_exposed": kw.pop("public_exposed", False), "exposure_basis": kw.pop("exposure_basis", []),
        "private_ips": kw.pop("private_ips", []), "public_ips": kw.pop("public_ips", []),
        "description_purpose": kw.pop("description_purpose", ""), "environment": kw.pop("environment", ""),
    }
    base.update(kw)
    return base


def _document():
    # Topology: app instance (sg-app) in subnet-a -> RDS seed (sg-db) in subnet-b, same VPC.
    # SG-db ingress allows 5432 from sg-app. A "marketing" instance (sg-mkt) also
    # has an ingress path to the DB it should NOT (segmentation finding via out-of-scope decl).
    resources = [
        _res("arn:db", "payments-db", "rds:db-cluster", name="payments-db",
             relationships={"vpc": ["vpc-1"], "security_groups": ["sg-db"], "subnet_group": ["sg-db"]},
             tags={"pci:cde": "true", "data-classification": "chd"}),
        # DB ENI
        _res("arn:eni-db", "eni-db", "ec2:network-interface",
             relationships={"vpc": ["vpc-1"], "subnet": ["subnet-b"], "security_groups": ["sg-db"],
                            "attached_to": ["payments-db"]},
             private_ips=["10.0.2.10"]),
        # App instance + ENI
        _res("arn:app", "i-app", "ec2:instance", name="orders-app",
             relationships={"vpc": ["vpc-1"], "security_groups": ["sg-app"], "subnet": ["subnet-a"],
                            "enis": ["eni-app"]},
             private_ips=["10.0.1.10"]),
        _res("arn:eni-app", "eni-app", "ec2:network-interface",
             relationships={"vpc": ["vpc-1"], "subnet": ["subnet-a"], "security_groups": ["sg-app"],
                            "attached_to": ["i-app"]},
             private_ips=["10.0.1.10"]),
        # Marketing instance declared out-of-scope but with a path to the DB.
        _res("arn:mkt", "i-mkt", "ec2:instance", name="marketing-web",
             relationships={"vpc": ["vpc-1"], "security_groups": ["sg-mkt"], "subnet": ["subnet-a"],
                            "enis": ["eni-mkt"]},
             private_ips=["10.0.1.20"], tags={"pci:scope": "out"}),
        _res("arn:eni-mkt", "eni-mkt", "ec2:network-interface",
             relationships={"vpc": ["vpc-1"], "subnet": ["subnet-a"], "security_groups": ["sg-mkt"],
                            "attached_to": ["i-mkt"]},
             private_ips=["10.0.1.20"]),
        # Security groups: db allows 5432 from sg-app AND sg-mkt. Each SG carries
        # AWS's explicit default allow-all egress (as the real artifact does).
        _res("arn:sg-db", "sg-db", "ec2:security-group", name="db-sg",
             relationships={"vpc": ["vpc-1"],
                            "ingress_rules": ["ingress 6:5432 sg:sg-app", "ingress 6:5432 sg:sg-mkt"],
                            "egress_rules": ["egress -1:all 0.0.0.0/0"]}),
        _res("arn:sg-app", "sg-app", "ec2:security-group", name="app-sg",
             relationships={"vpc": ["vpc-1"], "ingress_rules": [],
                            "egress_rules": ["egress -1:all 0.0.0.0/0"]}),
        _res("arn:sg-mkt", "sg-mkt", "ec2:security-group", name="mkt-sg",
             relationships={"vpc": ["vpc-1"], "ingress_rules": [],
                            "egress_rules": ["egress -1:all 0.0.0.0/0"]}),
        # An IAM role that can read the CDE (no network path) -> security-impacting.
        _res("arn:role-etl", "etl-role", "iam:role", name="etl-role", region="GLOBAL",
             iam_policy_data={"principal_type": "iam:role",
                              "inline_policies": {"etl": {"Statement": [
                                  {"Effect": "Allow", "Action": "rds-db:connect", "Resource": "arn:db"}]}},
                              "trust_principals": ["arn:user-dev"]}),
        _res("arn:user-dev", "dev", "iam:user", name="dev", region="GLOBAL",
             iam_policy_data={"principal_type": "iam:user"}),
        # CloudTrail -> always security-impacting infra.
        _res("arn:trail", "main-trail", "cloudtrail:trail", name="main-trail"),
        # An isolated bucket (no path, no IAM) -> out-of-scope.
        _res("arn:s3-mkt", "marketing-assets", "s3:bucket", name="marketing-assets"),
    ]
    return {"schema_version": "1.1.0", "resources": resources,
            "regions_coverage": [{"account_id": "111122223333", "region": "us-east-1", "status": "included"}]}


def _netdata():
    """Same-VPC topology: route is 'local'; NACLs allow all. Mark fetched=True."""
    nd = NetworkData(fetched=True)
    allow_all_in = [NaclRule(100, "-1", False, "allow", "0.0.0.0/0", None, None)]
    allow_all_out = [NaclRule(100, "-1", True, "allow", "0.0.0.0/0", None, None)]
    for sid in ("subnet-a", "subnet-b"):
        nd.nacls_by_subnet[sid] = Nacl(f"acl-{sid}", "vpc-1", True, [sid],
                                       list(allow_all_in), list(allow_all_out))
        nd.route_tables_by_subnet[sid] = RouteTable(f"rt-{sid}", "vpc-1", [sid], True,
                                                    [Route("10.0.0.0/16", "", "local", "active")])
    return nd


def _run(seed_cfg):
    index = InventoryIndex(_document())
    resolver = SeedResolver(seed_cfg, index.resources)
    graph = ReachabilityGraph(index, _netdata())
    # expand from CDE seed endpoints
    seed_eps = []
    for arn in resolver.cde_arns:
        seed_eps += graph.endpoints_for_resource(index.get(arn))
    graph.expand_from_seeds(seed_eps)
    graph.finalize()
    iam = IamGraph(index)
    classifier = ScopeClassifier(index, resolver, graph, iam)
    return classifier.classify()


def test_seed_tag_makes_cde():
    cfg = load_scope_config(None)  # rely on tags in the doc
    result = _run(cfg)
    db = result.classifications["arn:db"]
    assert db.category == Category.CDE
    assert db.confidence == Confidence.DETERMINED
    assert any("seed:" in b for b in db.basis)


def test_seeds_file_with_empty_keys_parses(tmp_path):
    # A half-filled template — keys present but empty (commented-out entries) —
    # parses to None per-key in YAML; the loader must coerce None -> [] not crash.
    p = tmp_path / "seeds.yaml"
    p.write_text(
        "cde_resources:\n"
        "cde_networks:\n"
        "  vpcs:\n"
        "  subnets:\n"
        "  cidrs:\n"
        "out_of_scope_declared:\n",
        encoding="utf-8",
    )
    cfg = load_scope_config(str(p))
    assert cfg.cde_resources == []
    assert cfg.cde_vpcs == [] and cfg.cde_subnets == [] and cfg.cde_cidrs == []
    assert cfg.out_of_scope_declared == []
    assert cfg.has_any_seed is False


def test_seeds_file_populated_with_empty_subkey(tmp_path):
    # Real values flow through, and an empty sub-key (subnets:) coerces to [].
    p = tmp_path / "seeds.yaml"
    p.write_text(
        "cde_resources:\n"
        "  - my-cardholder-bucket\n"
        "cde_networks:\n"
        "  vpcs: [vpc-0cde111]\n"
        "  subnets:\n"
        "  cidrs: ['10.20.0.0/16']\n",
        encoding="utf-8",
    )
    cfg = load_scope_config(str(p))
    assert cfg.cde_resources == ["my-cardholder-bucket"]
    assert cfg.cde_vpcs == ["vpc-0cde111"]
    assert cfg.cde_subnets == []
    assert cfg.cde_cidrs == ["10.20.0.0/16"]
    assert cfg.has_any_seed is True


def test_reachability_connects_app_to_db():
    cfg = load_scope_config(None)
    result = _run(cfg)
    # The app instance reaches the DB on 5432 → connected-to, DETERMINED.
    app = result.classifications["arn:app"]
    assert app.category == Category.CONNECTED
    assert app.confidence == Confidence.DETERMINED
    # A proven path exists with the port.
    assert any("5432" in h.port for p in result.paths for h in p.hops)


def test_iam_security_impacting():
    cfg = load_scope_config(None)
    result = _run(cfg)
    etl = result.classifications["arn:role-etl"]
    assert etl.category == Category.SECURITY_IMPACTING
    # Assume-chain pulls in the dev user too.
    dev = result.classifications["arn:user-dev"]
    assert dev.category == Category.SECURITY_IMPACTING


def test_cloudtrail_always_security_impacting():
    cfg = load_scope_config(None)
    result = _run(cfg)
    assert result.classifications["arn:trail"].category == Category.SECURITY_IMPACTING


def test_segmentation_finding_for_declared_out():
    cfg = load_scope_config(None)
    result = _run(cfg)
    # marketing-web is tagged pci:scope=out but reaches the DB → finding.
    arns = {f.resource_arn for f in result.segmentation_findings}
    assert "arn:mkt" in arns


def test_isolated_bucket_out_of_scope():
    cfg = load_scope_config(None)
    result = _run(cfg)
    s3 = result.classifications["arn:s3-mkt"]
    # No path, no IAM → out-of-scope, but flagged as data-store candidate too.
    assert s3.category in (Category.OUT_OF_SCOPE, Category.UNDETERMINED)
    assert any("isolation supported" in n for n in s3.notes) or any("candidate-chd" in b for b in s3.basis)


def test_scoped_document_and_workbook(tmp_path):
    from openpyxl import load_workbook
    from pci_inventory.scope.runner import build_scoped_document
    from pci_inventory.scope.workbook import write_scope_workbook

    index = InventoryIndex(_document())
    resolver = SeedResolver(load_scope_config(None), index.resources)
    graph = ReachabilityGraph(index, _netdata())
    seed_eps = []
    for arn in resolver.cde_arns:
        seed_eps += graph.endpoints_for_resource(index.get(arn))
    graph.expand_from_seeds(seed_eps)
    graph.finalize()
    iam = IamGraph(index)
    result = ScopeClassifier(index, resolver, graph, iam).classify()

    # Scoped JSON document: superset of inventory.json + scope block.
    doc = build_scoped_document(index, result, graph, {"gap_fetched": True, "gap_notes": []})
    assert doc["schema_version"] == "1.1.0"  # Stage 1 data preserved
    assert "scope_analysis" in doc
    db = next(r for r in doc["resources"] if r["arn"] == "arn:db")
    assert db["pci_scope"] == "CDE"
    assert db["scope"]["confidence"] == "DETERMINED"
    assert doc["scope_analysis"]["segmentation_findings"]
    assert doc["scope_analysis"]["reachability_paths"]

    # Workbook: the four Stage-2 sheets exist.
    path = write_scope_workbook(index, result, graph, tmp_path / "scope.xlsx")
    wb = load_workbook(path)
    for sheet in ("Scope Classification", "Reachability Paths", "Segmentation Findings",
                  "IAM-to-CDE Access", "Scope Caveats"):
        assert sheet in wb.sheetnames


def test_seed_precedence_config_over_tag():
    # Tag says CDE, but config declares the same resource out-of-scope → config wins.
    doc = _document()
    cfg = load_scope_config(None)
    cfg.out_of_scope_declared = ["arn:db"]  # contradicts the pci:cde=true tag
    index = InventoryIndex(doc)
    resolver = SeedResolver(cfg, index.resources)
    seed = resolver.seed_for("arn:db")
    assert seed is not None and seed.kind.value == "out_of_scope_declared"
    assert resolver.conflicts  # the disagreement is recorded


def test_no_seed_mode_flags_only():
    # No tags recognized as seeds → strip them by using an empty resolver config
    # against a doc with no seed tags.
    doc = _document()
    for r in doc["resources"]:
        r["tags"] = {}
    index = InventoryIndex(doc)
    cfg = load_scope_config(None)
    resolver = SeedResolver(cfg, index.resources)
    graph = ReachabilityGraph(index, _netdata())
    iam = IamGraph(index)
    result = ScopeClassifier(index, resolver, graph, iam).classify()
    assert result.no_seed_mode is True
    # No resource may be asserted CDE/DETERMINED-connected.
    assert all(c.category != Category.CDE for c in result.classifications.values())
    assert any("NO SEEDS" in cav for cav in result.caveats)
