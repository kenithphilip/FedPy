"""Stage 2 workbook sheets, appended to the Stage 1 formatting standards.

Adds: Scope Classification, Reachability Paths, Segmentation Findings, and
IAM-to-CDE Access — plus a Scope Caveats sheet. Reuses Stage 1's palette and
header/freeze/autofilter helpers so formatting is consistent.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pci_inventory.output.workbook import (
    GOOD_FILL,
    RISK_FILL,
    RISK_FONT,
    SUBTITLE_FONT,
    TITLE_FONT,
    WARN_FILL,
    WRAP,
    _autofilter_and_freeze,
    _style_header,
)
from pci_inventory.scope.classifier import ScopeResult
from pci_inventory.scope.reachability import ReachabilityGraph

# Category cell fills for at-a-glance scanning.
_CATEGORY_FILL = {
    "CDE": RISK_FILL,                 # red — highest attention
    "connected-to": WARN_FILL,        # amber
    "security-impacting": WARN_FILL,  # amber
    "out-of-scope": GOOD_FILL,        # green
    "undetermined": None,
}


def _banner(ws: Worksheet, result: ScopeResult) -> int:
    """Write the no-seed / caveat banner at the top; return next free row."""
    row = 1
    if result.no_seed_mode:
        c = ws.cell(row=row, column=1,
                    value="⚠ NO SEEDS PROVIDED — no in-scope determination was made. "
                          "Only heuristic candidates are flagged. Declare seeds "
                          "(docs/scope-seed-and-tagging-convention.md) for a real analysis.")
        c.font = Font(bold=True, color="9C0006", size=12)
        c.fill = RISK_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2
    return row


def _build_scope_classification(ws: Worksheet, index, result: ScopeResult) -> None:
    row = _banner(ws, result)
    ws.cell(row=row, column=1, value="Scope Classification — every resource with basis + confidence").font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1,
            value="Categories: CDE / connected-to / security-impacting / out-of-scope / undetermined. "
                  "Confidence: DETERMINED (seed / proven path / IAM grant), CANDIDATE (heuristic), UNDETERMINED. "
                  "Out-of-scope = no path/relationship found; it is NOT proof of absence of CHD.").alignment = WRAP
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 2

    header_row = row
    headers = ["ARN", "Resource Type", "Name", "Region", "Category (primary)",
               "Also (secondary)", "Confidence", "Basis", "Notes"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, len(headers), row=header_row)

    r = header_row + 1
    # Stable order: category importance, then arn.
    order = {"CDE": 0, "connected-to": 1, "security-impacting": 2, "undetermined": 3, "out-of-scope": 4}
    rows = []
    for res in index.resources:
        arn = res.get("arn", "")
        cl = result.classifications.get(arn)
        if cl is None:
            continue
        rows.append((order.get(cl.category.value, 9), arn, res, cl))
    for _, arn, res, cl in sorted(rows, key=lambda x: (x[0], x[1])):
        ws.cell(row=r, column=1, value=arn)
        ws.cell(row=r, column=2, value=res.get("resource_type", ""))
        ws.cell(row=r, column=3, value=res.get("name", ""))
        ws.cell(row=r, column=4, value=res.get("region", ""))
        cat_cell = ws.cell(row=r, column=5, value=cl.category.value)
        fill = _CATEGORY_FILL.get(cl.category.value)
        if fill:
            cat_cell.fill = fill
            if cl.category.value == "CDE":
                cat_cell.font = RISK_FONT
        sec = ws.cell(row=r, column=6, value=", ".join(cl.secondary_categories()))
        if cl.secondary_categories():
            sec.fill = WARN_FILL  # multiple in-scope reasons — don't miss the others
        ws.cell(row=r, column=7, value=cl.confidence.value)
        bc = ws.cell(row=r, column=8, value="; ".join(cl.basis))
        bc.alignment = WRAP
        nc = ws.cell(row=r, column=9, value="; ".join(cl.notes))
        nc.alignment = WRAP
        r += 1

    widths = [46, 26, 24, 12, 18, 22, 13, 50, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _autofilter_and_freeze(ws, len(headers), ws.max_row, header_row=header_row)


def _build_reachability_paths(ws: Worksheet, graph: ReachabilityGraph) -> None:
    ws.cell(row=1, column=1, value="Reachability Paths — provable permitted paths (route ∧ SG ∧ NACL)").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Each row is one proven hop placing a resource in scope: the source, the CDE "
                  "target, protocol/port, the routing path, the permitting SG rule, and the NACL "
                  "evaluation. Confidence is CANDIDATE when live NACL/route data was unavailable.").alignment = WRAP
    ws.merge_cells("A2:H2")
    header_row = 4
    headers = ["Path ID", "Direction", "Source", "Target (CDE)", "Proto:Port",
               "Via", "Permitting SG Rule", "NACL Evaluation"]
    for i, head in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=head)
    _style_header(ws, len(headers), row=header_row)
    r = header_row + 1
    if not graph.paths:
        ws.cell(row=r, column=1, value="(no permitted paths to/from seeds found)")
        r += 1
    for p in graph.paths:
        for h in p.hops:
            ws.cell(row=r, column=1, value=p.path_id)
            ws.cell(row=r, column=2, value=p.direction)
            ws.cell(row=r, column=3, value=h.src)
            ws.cell(row=r, column=4, value=h.dst)
            ws.cell(row=r, column=5, value=f"{h.proto}:{h.port}")
            ws.cell(row=r, column=6, value=h.via)
            ws.cell(row=r, column=7, value=h.sg_rule)
            ws.cell(row=r, column=8, value=h.nacl_note)
            if p.confidence == "CANDIDATE":
                ws.cell(row=r, column=2).fill = WARN_FILL
            r += 1
    widths = [12, 12, 40, 40, 14, 22, 36, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _autofilter_and_freeze(ws, len(headers), ws.max_row, header_row=header_row)


def _build_segmentation_findings(ws: Worksheet, result: ScopeResult) -> None:
    ws.cell(row=1, column=1, value="Segmentation Findings — out-of-scope resources that touch the CDE").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="A FINDING = an out-of-scope resource (human-declared OR tool-derived) that has a "
                  "permitted relationship to the CDE. Kind: INBOUND (it can reach into the CDE — the "
                  "primary segmentation-failure case), OUTBOUND (the CDE reaches it — often expected), "
                  "or IAM (a control-plane path). Absence of findings is NOT proof of complete "
                  "segmentation — penetration testing (Req 11.4.x) is required.").alignment = WRAP
    ws.merge_cells("A2:G2")
    header_row = 4
    headers = ["Severity", "Kind", "Declared?", "Resource", "Expected", "Path ID", "Offending Path / Relationship"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, len(headers), row=header_row)
    r = header_row + 1
    if not result.segmentation_findings:
        ws.cell(row=r, column=1,
                value="(no out-of-scope→CDE relationships found — NOT proof of complete segmentation; "
                      "Req 11.4.x pen-test still required)")
        r += 1
    for f in result.segmentation_findings:
        sev_cell = ws.cell(row=r, column=1, value=f.severity)
        # Inbound contradictions are the audit-critical ones → red; outbound → amber.
        if f.kind == "outbound":
            sev_cell.fill = WARN_FILL
        else:
            sev_cell.fill = RISK_FILL
            sev_cell.font = RISK_FONT
        ws.cell(row=r, column=2, value=f.kind)
        ws.cell(row=r, column=3, value="Yes" if f.declared else "no (tool-derived)")
        ws.cell(row=r, column=4, value=f.resource_arn)
        ws.cell(row=r, column=5, value=f.expected)
        ws.cell(row=r, column=6, value=f.path_id)
        pc = ws.cell(row=r, column=7, value=f.path_summary)
        pc.alignment = WRAP
        r += 1
    widths = [10, 10, 16, 44, 28, 12, 64]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _autofilter_and_freeze(ws, len(headers), ws.max_row, header_row=header_row)


def _build_iam_to_cde(ws: Worksheet, result: ScopeResult) -> None:
    ws.cell(row=1, column=1, value="IAM-to-CDE Access — principals that can act on the CDE").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Static over-approximation: identity + resource policies + assume-role chain, "
                  "intersected with the CDE set. Does NOT resolve SCPs / permission boundaries / "
                  "condition keys — flags candidate access with the granting statement, not "
                  "effective access.").alignment = WRAP
    ws.merge_cells("A2:F2")
    header_row = 4
    headers = ["Principal", "Principal Type", "CDE Resource", "Capability", "Via", "Confidence"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, len(headers), row=header_row)
    r = header_row + 1
    if not result.iam_findings:
        ws.cell(row=r, column=1, value="(no IAM principals with CDE-affecting access found)")
        r += 1
    for f in result.iam_findings:
        ws.cell(row=r, column=1, value=f.principal_arn)
        ws.cell(row=r, column=2, value=f.principal_type)
        ws.cell(row=r, column=3, value=f.cde_resource_arn)
        cc = ws.cell(row=r, column=4, value=f.capability)
        if f.capability in ("full-cde-bucket", "full-cde-key", "full-cde-db", "all-actions"):
            cc.fill = RISK_FILL
        ws.cell(row=r, column=5, value=f.via)
        ws.cell(row=r, column=6, value=f.confidence)
        r += 1
    widths = [46, 22, 46, 24, 30, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _autofilter_and_freeze(ws, len(headers), ws.max_row, header_row=header_row)


def _build_caveats(ws: Worksheet, result: ScopeResult) -> None:
    ws.cell(row=1, column=1, value="Scope Analysis — Caveats & Method").font = TITLE_FONT
    r = 3
    for cav in result.caveats:
        c = ws.cell(row=r, column=1, value=f"• {cav}")
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
    if result.seed_conflicts:
        r += 1
        ws.cell(row=r, column=1, value="Seed precedence conflicts (config won):").font = SUBTITLE_FONT
        r += 1
        for conflict in result.seed_conflicts:
            ws.cell(row=r, column=1, value=f"• {conflict}").alignment = WRAP
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1
    ws.column_dimensions["A"].width = 120


def append_scope_sheets(wb: Workbook, index, result: ScopeResult,
                        graph: ReachabilityGraph) -> None:
    """Append the four Stage-2 sheets + caveats to an existing workbook."""
    _build_scope_classification(wb.create_sheet("Scope Classification"), index, result)
    _build_reachability_paths(wb.create_sheet("Reachability Paths"), graph)
    _build_segmentation_findings(wb.create_sheet("Segmentation Findings"), result)
    _build_iam_to_cde(wb.create_sheet("IAM-to-CDE Access"), result)
    _build_caveats(wb.create_sheet("Scope Caveats"), result)


def write_scope_workbook(index, result: ScopeResult, graph: ReachabilityGraph,
                         path: str | Path) -> Path:
    """Write a standalone scope workbook (Stage-2 sheets only)."""
    wb = Workbook()
    wb.remove(wb.active)
    append_scope_sheets(wb, index, result, graph)
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
