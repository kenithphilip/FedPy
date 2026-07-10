"""Stage 3 workbook — the final consolidated QSA deliverable.

Appends, to the Stage 1 + Stage 2 sheets, one Evidence sheet per PCI requirement
domain (Req 01–12), a PCI Requirement Mapping sheet, a Findings & Indicators
sheet, and a QSA Notes sheet. Reuses Stage 1 formatting (frozen headers,
autofilters, conditional highlighting). Indicators are loudly labelled as aids,
not determinations.
"""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from pci_inventory.evidence.mapping import COLUMN_REQUIREMENT_MAP
from pci_inventory.evidence.models import (
    RequirementDomain,
    sheet_title,
)
from pci_inventory.evidence.runner import EvidenceResult
from pci_inventory.output.workbook import (
    GOOD_FILL,
    RISK_FILL,
    RISK_FONT,
    SUBTITLE_FONT,
    TITLE_FONT,
    WARN_FILL,
    WRAP,
    _autofilter_and_freeze,
    _style_header,
)

_CAT_FILL = {
    "CDE": RISK_FILL, "connected-to": WARN_FILL, "security-impacting": WARN_FILL,
    "out-of-scope": GOOD_FILL, "undetermined": None,
}
# A few field values worth highlighting red across requirement sheets.
_RED_FIELD_VALUES = {"No", "NON_COMPLIANT", "no-scan-on-push"}


def _build_requirement_sheet(ws: Worksheet, domain: RequirementDomain,
                             rows: list, scope_missing: bool) -> None:
    ws.cell(row=1, column=1, value=domain.value).font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Each row = one resource's already-collected evidence for this requirement, "
                  "traceable to its ARN and carrying the Stage 2 scope classification. "
                  "Indicators/values here assist assessment — they are NOT compliance "
                  "determinations.").alignment = WRAP
    ws.merge_cells("A2:H2")

    if not rows:
        ws.cell(row=4, column=1, value="(no resources produced evidence for this requirement — "
                                       "see QSA Notes for shared-responsibility / not-collectable items)")
        return

    # Union of evidence field keys (stable order by first appearance).
    field_keys: list[str] = []
    for row in rows:
        for k in row.fields:
            if k not in field_keys:
                field_keys.append(k)

    header_row = 4
    headers = (["Resource ARN", "Name", "Type", "Region", "Scope", "Confidence", "Sub-reqs"]
               + field_keys + ["Findings", "Notes"])
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, len(headers), row=header_row)

    r = header_row + 1
    for row in sorted(rows, key=lambda x: (_scope_rank(x.scope_category), x.resource_arn)):
        ws.cell(row=r, column=1, value=row.resource_arn)
        ws.cell(row=r, column=2, value=row.name)
        ws.cell(row=r, column=3, value=row.resource_type)
        ws.cell(row=r, column=4, value=row.region)
        sc = ws.cell(row=r, column=5, value=row.scope_category)
        fill = _CAT_FILL.get(row.scope_category)
        if fill:
            sc.fill = fill
            if row.scope_category == "CDE":
                sc.font = RISK_FONT
        ws.cell(row=r, column=6, value=row.scope_confidence)
        ws.cell(row=r, column=7, value=row.sub_requirements)
        for j, k in enumerate(field_keys, start=8):
            val = row.fields.get(k, "")
            cell = ws.cell(row=r, column=j, value=val if not isinstance(val, (list, dict)) else str(val))
            if str(val) in _RED_FIELD_VALUES:
                cell.fill = RISK_FILL
                cell.font = RISK_FONT
        fcell = ws.cell(row=r, column=8 + len(field_keys), value="\n".join(row.findings))
        fcell.alignment = WRAP
        if row.findings:
            fcell.fill = WARN_FILL
        ws.cell(row=r, column=9 + len(field_keys), value=row.notes).alignment = WRAP
        r += 1

    # Reasonable widths.
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    for j in range(8, 8 + len(field_keys)):
        ws.column_dimensions[get_column_letter(j)].width = 20
    ws.column_dimensions[get_column_letter(8 + len(field_keys))].width = 40
    _autofilter_and_freeze(ws, len(headers), ws.max_row, header_row=header_row)


def _scope_rank(category: str) -> int:
    return {"CDE": 0, "connected-to": 1, "security-impacting": 2,
            "undetermined": 3, "out-of-scope": 4}.get(category, 9)


def _build_mapping_sheet(ws: Worksheet) -> None:
    ws.cell(row=1, column=1, value="PCI Requirement Mapping — data point → requirement(s)").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Every inventory/evidence data point across the whole tool (Stages 1–3) and the "
                  "PCI DSS v4.0.1 requirement(s) it supports. Source: INV = collected at inventory "
                  "time; FOLLOW-UP = Stage 3 targeted query; DERIVED = computed indicator; "
                  "SHARED = AWS shared-responsibility; NOT_COLLECTABLE = not observable read-only.").alignment = WRAP
    ws.merge_cells("A2:C2")
    header_row = 4
    headers = ["Data Point / Column", "PCI Requirement(s)", "Source"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, len(headers), row=header_row)
    r = header_row + 1
    for dp, req, src in COLUMN_REQUIREMENT_MAP:
        ws.cell(row=r, column=1, value=dp).alignment = WRAP
        ws.cell(row=r, column=2, value=req)
        ws.cell(row=r, column=3, value=src)
        r += 1
    for i, w in enumerate((58, 30, 36), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    _autofilter_and_freeze(ws, len(headers), ws.max_row, header_row=header_row)


def _build_indicators_sheet(ws: Worksheet, result: EvidenceResult) -> None:
    ind = result.indicators
    ws.cell(row=1, column=1, value="Findings & Indicators").font = TITLE_FONT
    banner = ws.cell(row=2, column=1, value="⚠ " + ind.disclaimer)
    banner.font = Font(bold=True, color="9C0006")
    banner.fill = WARN_FILL
    banner.alignment = WRAP
    ws.merge_cells("A2:F2")

    r = 4
    ws.cell(row=r, column=1, value="Coverage indicators — overall and by scope category").font = SUBTITLE_FONT
    r += 1
    scopes = ["overall"] + list(ind.by_scope.keys())
    metric_keys = list(ind.overall.keys())
    ws.cell(row=r, column=1, value="Metric").font = Font(bold=True)
    for c, s in enumerate(scopes, start=2):
        ws.cell(row=r, column=c, value=s).font = Font(bold=True)
    _style_header(ws, 1 + len(scopes), row=r)
    r += 1
    for mk in metric_keys:
        ws.cell(row=r, column=1, value=mk)
        for c, s in enumerate(scopes, start=2):
            block = ind.overall if s == "overall" else ind.by_scope.get(s, {})
            ws.cell(row=r, column=c, value=block.get(mk, ""))
        r += 1

    # Risk lists.
    r += 1
    ws.cell(row=r, column=1, value="Risk lists (indicators warranting review)").font = SUBTITLE_FONT
    r += 1
    for name, items in ind.risk_lists.items():
        label = ws.cell(row=r, column=1, value=f"{name} ({len(items)})")
        label.font = Font(bold=True)
        if items:
            label.fill = RISK_FILL if name in ("public_exposed", "unencrypted_at_rest",
                                                "overly_permissive_iam", "publicly_shared") else WARN_FILL
        r += 1
        for item in items[:200]:  # cap per list in the sheet; full set in JSON
            ws.cell(row=r, column=2, value=item).alignment = WRAP
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
            r += 1
        if len(items) > 200:
            ws.cell(row=r, column=2, value=f"… +{len(items) - 200} more (see inventory-evidence.json)")
            r += 1

    # Security-service findings rollup.
    r += 1
    ws.cell(row=r, column=1, value="Security-service findings (rollup)").font = SUBTITLE_FONT
    r += 1
    by_source: dict[str, int] = defaultdict(int)
    for f in result.findings.findings:
        by_source[f.source] += 1
    if not by_source:
        ws.cell(row=r, column=1, value="(no findings collected — services not enabled or no credentials)")
        r += 1
    for src, count in sorted(by_source.items()):
        ws.cell(row=r, column=1, value=src)
        ws.cell(row=r, column=2, value=count)
        r += 1

    ws.column_dimensions["A"].width = 40
    for c in range(2, 8):
        ws.column_dimensions[get_column_letter(c)].width = 18


def _build_qsa_notes(ws: Worksheet, result: EvidenceResult) -> None:
    ws.cell(row=1, column=1, value="QSA Notes — Scope Determination, Limitations & Attestation").font = TITLE_FONT
    r = 3
    ws.cell(row=r, column=1, value="How scope was determined").font = SUBTITLE_FONT
    r += 1
    how = [
        "Scope originates from human-declared SEEDS (resources/networks attested to store/process/"
        "transmit cardholder data) — config > tags (pci:cde / pci:scope / data-classification) > CLI flags.",
        "Stage 2 expands from seeds via a reachability graph (a permitted path requires route table "
        "AND security group AND network ACL to all allow it; multi-hop chains composed) and an IAM "
        "graph (principals that can act on a CDE resource), then validates segmentation (out-of-scope "
        "resources that can still reach the CDE are findings).",
        "Each resource carries a scope category (CDE / connected-to / security-impacting / "
        "out-of-scope / undetermined) with a confidence (DETERMINED / CANDIDATE / UNDETERMINED) and "
        "the basis. Without seeds, nothing is asserted in-scope.",
    ]
    r = _write_bullets(ws, r, how)
    r += 1
    ws.cell(row=r, column=1, value="Known limitations & shared-responsibility caveats").font = SUBTITLE_FONT
    r += 1
    r = _write_bullets(ws, r, result.caveats)
    ws.column_dimensions["A"].width = 130


def _write_bullets(ws: Worksheet, r: int, items: list[str]) -> int:
    for text in items:
        c = ws.cell(row=r, column=1, value=f"• {text}")
        c.alignment = WRAP
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
    return r


# Domain → builder order.
_DOMAIN_ORDER = list(RequirementDomain)


def append_evidence_sheets(wb: Workbook, result: EvidenceResult) -> None:
    """Append the Stage 3 sheets to a workbook."""
    rows_by_domain: dict[RequirementDomain, list] = defaultdict(list)
    for row in result.rows:
        rows_by_domain[row.requirement].append(row)
    for domain in _DOMAIN_ORDER:
        ws = wb.create_sheet(sheet_title(domain))
        _build_requirement_sheet(ws, domain, rows_by_domain.get(domain, []), result.scope_missing)
    _build_mapping_sheet(wb.create_sheet("PCI Requirement Mapping"))
    _build_indicators_sheet(wb.create_sheet("Findings & Indicators"), result)
    _build_qsa_notes(wb.create_sheet("QSA Notes"), result)


def write_evidence_workbook(result: EvidenceResult, path: str | Path) -> Path:
    """Write a standalone evidence workbook (Stage 3 sheets only)."""
    wb = Workbook()
    wb.remove(wb.active)
    append_evidence_sheets(wb, result)
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out
